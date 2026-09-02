"""Construye el instrumento definitivo de adjudicacion (ds04_pool.xlsx),
parametrizado por numero de consultas (MATCHING-SPEC.md §4.4.1).

NO SE HA EJECUTADO TODAVIA. El piloto (docs/adjudicacion-piloto*.xlsx) midio
un ritmo real de 6-9 min/par, muy por encima del "uno o dos minutos por fila"
que asumia el instrumento del piloto. A 206 pares eso son entre ~21 y ~31
horas (3 min/par optimista: ~10 h) -- ver MATCHING-SPEC.md §4.4.1. Generar
el instrumento definitivo esta bloqueado hasta que el adjudicador confirme
cuanta disponibilidad real tiene; ese numero decide --n-queries.

Reutiliza el pool ya construido por build_test_collection.py (mismas 20
consultas, mismo corpus, misma logica de ranking) -- no reimplementa nada de
eso. Lo nuevo aqui es: (1) reducir el numero de consultas sin romper el
requisito de >=3 consultas de especialidad escasa, (2) ensamblar el XLSX de
tres hojas con cegamiento, duplicados y justificaciones parametrizados, en
vez de solo reportar conteos como hace build_test_collection.py.

Cambios de redaccion en la hoja de instrucciones respecto del piloto
(ninguno toca la rubrica del protocolo, seccion 4):

  1. Se retira la promesa de "uno o dos minutos por fila" -- el piloto la
     desmintio por un factor de 3x-6x. Prometer un tiempo falso presiona al
     adjudicador a apurar el juicio.
  2. Se agrega una regla explicita sobre la via juridica: se toma como dada:
     si el adjudicador discrepa de ella, lo anota en la justificacion en vez
     de bajar la calificacion del abogado. Esta regla YA existe en el
     protocolo (protocolo-adjudicacion_1.docx §4.2) pero no era visible en
     el instrumento -- el piloto mostro el sintoma: 52% de las respuestas
     cayeron en la categoria intermedia (1), y varias justificaciones
     corregian el encuadre juridico del caso en vez de evaluar al abogado.
  3. Se mantiene sin cambios la frase sobre "confianza baja es una respuesta
     valida" -- el piloto la valido: se uso en 1 de 21 filas (~5%) y fue
     justo ese par el que produjo la unica inconsistencia intra-evaluador
     detectada entre los pares duplicados.

Uso (NO EJECUTAR sin decidir antes --n-queries con el adjudicador):
    cd ml-service
    pip install -r evaluation/requirements-eval.txt
    python evaluation/build_instrument.py --n-queries 14 --duplicate-rate 0.12 \
        --out data/reference/ds04_pool.xlsx
"""

from __future__ import annotations

import argparse
import json
import sys
from pathlib import Path
from random import Random

if sys.stdout.encoding and sys.stdout.encoding.lower() != "utf-8":
    sys.stdout.reconfigure(encoding="utf-8")

sys.path.insert(0, str(Path(__file__).resolve().parent))
sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from openpyxl import Workbook  # noqa: E402
from openpyxl.styles import Alignment, Font, PatternFill  # noqa: E402
from openpyxl.worksheet.datavalidation import DataValidation  # noqa: E402

from build_test_collection import (  # noqa: E402
    CORPUS_PATH,
    POOL_DEPTH,
    QUERIES,
    RANDOM_STATE,
    SCARCE_AREAS,
    pool_for_query,
)

MIN_N_QUERIES = 8  # por debajo de esto no hay margen para cubrir escasas + variedad de urgencia/complejidad
MAX_N_QUERIES = len(QUERIES)  # 20
MIN_SCARCE_REQUIRED = 3  # igual que build_test_collection.py -- no se relaja al reducir N

DEFAULT_DUPLICATE_RATE = 0.15  # > 10 % -- ver hallazgo del piloto, antes 0.10
DEFAULT_JUSTIFICATION_RATE = 0.15  # protocolo §5.4 / §6

YELLOW = PatternFill(start_color="FFFFF2CC", end_color="FFFFF2CC", fill_type="solid")

# ═══════════════════════════════════════════════════════════════════════════
# Orden de prioridad para reducir el numero de consultas.
#
# Curado a mano UNA sola vez (no aleatorio, no recalculado en cada corrida):
# intercala las 5 consultas de especialidad escasa (q16-q20) con consultas de
# cobertura alta/media variando urgencia y complejidad, de modo que tomar los
# primeros N de esta lista para cualquier N >= 12 conserve las 5 escasas
# completas, y para N >= 9 conserve al menos 3 (el minimo exigido). No es una
# muestra aleatoria estratificada -- es una eleccion fija, documentada aqui,
# para que --n-queries sea reproducible sin depender de una semilla.
# ═══════════════════════════════════════════════════════════════════════════
QUERY_PRIORITY_ORDER = [
    "q16", "q01", "q17", "q02", "q18", "q04", "q19", "q07", "q20", "q10",
    "q03", "q05", "q06", "q08", "q09", "q11", "q12", "q13", "q14", "q15",
]
assert set(QUERY_PRIORITY_ORDER) == {q["query_id"] for q in QUERIES}, (
    "QUERY_PRIORITY_ORDER debe cubrir exactamente las 20 query_id de QUERIES"
)


def select_queries(n: int) -> list[dict]:
    if not (MIN_N_QUERIES <= n <= MAX_N_QUERIES):
        raise SystemExit(f"--n-queries debe estar entre {MIN_N_QUERIES} y {MAX_N_QUERIES}, recibido {n}")
    by_id = {q["query_id"]: q for q in QUERIES}
    chosen_ids = set(QUERY_PRIORITY_ORDER[:n])
    scarce_chosen = sum(1 for qid in chosen_ids if by_id[qid]["medical_specialty"] in SCARCE_AREAS)
    if scarce_chosen < MIN_SCARCE_REQUIRED:
        raise SystemExit(
            f"Con --n-queries {n} solo quedan {scarce_chosen} consultas de área escasa "
            f"(mínimo {MIN_SCARCE_REQUIRED}). Ajusta QUERY_PRIORITY_ORDER o sube --n-queries."
        )
    # se devuelven en el orden original (q01..q20), no en el de prioridad --
    # el orden de prioridad es solo el criterio de selección, no el de presentación
    return [q for q in QUERIES if q["query_id"] in chosen_ids]


def min_duplicate_gap(total_rows: int) -> int:
    """Separación mínima entre un duplicado y su primera aparición.

    El protocolo (§5.3) fija 30 posiciones para un instrumento de ~200+
    filas. Con --n-queries reducido el instrumento puede tener muchas menos
    filas que 30, y la regla literal se vuelve inaplicable. Regla adoptada
    aquí, documentada como desviación explícita (no silenciosa): al menos
    30 posiciones, o un cuarto del total de filas si el instrumento es más
    chico que 120 filas -- lo que sea menor, con un piso absoluto de 5.
    """
    return max(5, min(30, total_rows // 4))


def build_rows(
    queries: list[dict],
    lawyers: list[dict],
    rng: Random,
    duplicate_rate: float,
    justification_rate: float,
) -> list[dict]:
    lawyer_by_id = {l["lawyer_id"]: l for l in lawyers}

    base_rows: list[dict] = []
    for idx, q in enumerate(queries):
        variant_lists = pool_for_query(q, lawyers, idx)
        union = sorted({lid for ids in variant_lists.values() for lid in ids})
        rng.shuffle(union)  # cegamiento: orden de candidatos aleatorio, no por score
        for lid in union:
            l = lawyer_by_id[lid]
            base_rows.append({
                "case_description": q["case_description"],
                "medical_specialty": q["medical_specialty"],
                "perceived_urgency": q["perceived_urgency"],
                "lawyer_id": lid,
                "full_name": l["full_name"],
                "specialties": ", ".join(l["specialties"]),
                "medical_areas": ", ".join(l["medical_areas"]),
                "years_experience": l["years_experience"],
            })
    rng.shuffle(base_rows)  # cegamiento adicional: no agrupar visiblemente por consulta

    n_base = len(base_rows)
    n_duplicates = round(n_base * duplicate_rate)
    gap = min_duplicate_gap(n_base + n_duplicates)

    rows = list(base_rows)
    if n_duplicates > 0:
        candidates = list(range(n_base))
        rng.shuffle(candidates)
        chosen_originals = candidates[:n_duplicates]
        for orig_idx in chosen_originals:
            dup = dict(rows[orig_idx])
            # posición válida: al menos `gap` posiciones después del original,
            # dentro del largo actual de `rows` (que crece en cada inserción)
            min_pos = orig_idx + gap
            max_pos = len(rows)
            insert_at = rng.randint(min(min_pos, max_pos), max_pos)
            rows.insert(insert_at, dup)

    n_total = len(rows)
    n_justify = round(n_total * justification_rate)
    justify_idx = set(rng.sample(range(n_total), min(n_justify, n_total)))
    for i, row in enumerate(rows):
        row["justify"] = i in justify_idx

    return rows


INSTRUCCIONES_TEXT = [
    ("title", "Adjudicación de relevancia — instrumento definitivo"),
    ("subtitle", "Proyecto Sinapsistencia · Emparejamiento médico–abogado"),
    ("h", "Qué vas a hacer"),
    ("p", "En la hoja «Adjudicación» vas a ver pares. Cada fila tiene, a la izquierda, el "
          "resumen de una consulta médico-legal y, a la derecha, el perfil de un abogado."),
    ("p", "Para cada fila respondes una sola pregunta:"),
    ("q", "¿Qué tan apropiado es este abogado para atender esta consulta?"),
    ("h", "Cómo respondes"),
    ("p", "Llenas las columnas amarillas. Las demás son solo de lectura."),
    ("scale", "     2  =  Claramente apropiado. La vía legal que el caso necesita coincide con "
              "su especialidad, y además tiene experiencia en esa área médica o en una cercana."),
    ("scale", "     1  =  Parcialmente apropiado. Coincide una cosa pero no la otra. Podría "
              "llevar el caso, pero no es el perfil ideal."),
    ("scale", "     0  =  No apropiado. No es su vía legal, o el área médica le es ajena."),
    ("p", "Confianza: marcas Alta, Media o Baja según qué tan seguro te sientas del juicio."),
    ("p", "Marcar «Baja» es una respuesta perfectamente válida. Si el caso está fuera de tu "
          "terreno o la descripción no alcanza, márcalo así. Esa duda me sirve tanto como una "
          "respuesta segura."),
    ("p", "Justificación: solo en las filas donde diga «Sí» en la columna correspondiente. Una "
          "frase corta basta."),
    ("h", "Cosas importantes"),
    ("bullet", "Cada fila se juzga por separado. No estás eligiendo al mejor abogado ni "
               "ordenándolos."),
    ("bullet", "No juzgas el caso médico. No importa si hubo mala praxis ni qué tan grave fue. "
               "Solo si el abogado calza."),
    ("bullet", "La vía legal del caso (penal, civil, administrativa sanitaria, laboral-sanitaria "
               "o de seguros) se toma como dada — no es parte de lo que evalúas. Si no estás de "
               "acuerdo con esa clasificación, anótalo en la justificación en lugar de bajar la "
               "calificación del abogado: estás evaluando si el abogado calza con esa vía, no si "
               "la vía es la correcta."),
    ("bullet", "Ignora la disponibilidad o la carga de trabajo. El sistema ya filtra por eso "
               "antes."),
    ("bullet", "Ve a tu ritmo. Tu primera impresión profesional es lo que busco, pero no hay un "
               "tiempo objetivo por fila — trabaja en las tandas que te acomoden."),
    ("bullet", "Verás casos que se parecen entre sí. Es a propósito. No trates de recordar qué "
               "respondiste antes."),
    ("bullet", "No consultes nada externo: ni buscadores, ni herramientas de inteligencia "
               "artificial, ni colegas. Lo que necesito medir es tu criterio, el que traes de tu "
               "experiencia pericial. Si la respuesta viene de otra herramienta, la tesis pierde "
               "exactamente lo que la hace válida."),
    ("h", "Cuándo y cómo"),
    ("p", "Puedes hacerlo cuando te acomode y en las tandas que quieras. Solo te pido que anotes "
          "en la hoja «Registro» las fechas y los horarios en que trabajaste, y que al terminar "
          "completes la declaración que está al final de esa hoja."),
    ("p", "Si algo te resulta confuso o mal planteado, anótalo en la hoja «Registro»."),
    ("p", "Los datos de este ejercicio son simulados. No corresponden a pacientes, casos ni "
          "profesionales reales."),
]


def write_instrucciones(ws) -> None:
    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 110
    r = 2
    for kind, text in INSTRUCCIONES_TEXT:
        cell = ws.cell(row=r, column=2, value=text)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        if kind == "title":
            cell.font = Font(size=14, bold=True)
        elif kind in ("h",):
            cell.font = Font(size=12, bold=True)
            r += 1
        elif kind == "q":
            cell.font = Font(italic=True, bold=True)
        r += 2 if kind in ("title", "subtitle") else 1
    ws.row_dimensions[1].height = 4


def write_adjudicacion(ws, rows: list[dict]) -> None:
    headers = [
        "ID", "Resumen de la consulta", "Especialidad médica", "Urgencia",
        "Abogado", "Especialidad legal", "Áreas médicas", "Años",
        "Relevancia (0/1/2)", "Confianza", "¿Justificar?", "Justificación",
    ]
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = Font(bold=True)

    dv_relevance = DataValidation(type="list", formula1='"0,1,2"', allow_blank=True)
    dv_confidence = DataValidation(type="list", formula1='"Alta,Media,Baja"', allow_blank=True)
    ws.add_data_validation(dv_relevance)
    ws.add_data_validation(dv_confidence)

    for i, row in enumerate(rows, start=1):
        r = i + 1
        ws.cell(row=r, column=1, value=f"P{i:02d}" if len(rows) < 100 else f"P{i:03d}")
        ws.cell(row=r, column=2, value=row["case_description"])
        ws.cell(row=r, column=3, value=row["medical_specialty"])
        ws.cell(row=r, column=4, value=row["perceived_urgency"].capitalize())
        ws.cell(row=r, column=5, value=row["full_name"])
        ws.cell(row=r, column=6, value=row["specialties"])
        ws.cell(row=r, column=7, value=row["medical_areas"])
        ws.cell(row=r, column=8, value=row["years_experience"])
        rel_cell = ws.cell(row=r, column=9)
        conf_cell = ws.cell(row=r, column=10)
        rel_cell.fill = YELLOW
        conf_cell.fill = YELLOW
        dv_relevance.add(rel_cell)
        dv_confidence.add(conf_cell)
        if row["justify"]:
            ws.cell(row=r, column=11, value="Sí")
            just_cell = ws.cell(row=r, column=12)
            just_cell.fill = YELLOW

    ws.column_dimensions["B"].width = 60
    ws.column_dimensions["F"].width = 28
    ws.column_dimensions["G"].width = 30
    ws.column_dimensions["L"].width = 50
    for col_letter in ("A", "C", "D", "E", "H", "I", "J", "K"):
        ws.column_dimensions[col_letter].width = 16
    ws.freeze_panes = "A2"
    for r in range(2, len(rows) + 2):
        ws.cell(row=r, column=2).alignment = Alignment(wrap_text=True, vertical="top")


def write_registro(ws, n_rows: int) -> None:
    ws.column_dimensions["B"].width = 32
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 16

    ws.cell(row=2, column=2, value="Registro de la sesión").font = Font(size=13, bold=True)
    ws.cell(row=4, column=2, value="Anota las tandas en que trabajaste. No importa cuántas sean.")

    ws.cell(row=6, column=2, value="Tanda").font = Font(bold=True)
    ws.cell(row=6, column=3, value="Fecha").font = Font(bold=True)
    ws.cell(row=6, column=4, value="Hora de inicio").font = Font(bold=True)
    ws.cell(row=6, column=5, value="Hora de término").font = Font(bold=True)
    for i in range(8):  # más tandas en blanco que el piloto -- instrumento más largo
        r = 7 + i
        ws.cell(row=r, column=2, value=f"Tanda {i + 1}")
        for col in (3, 4, 5):
            ws.cell(row=r, column=col).fill = YELLOW

    r = 7 + 8 + 1
    ws.cell(row=r, column=2, value="Avance").font = Font(size=12, bold=True)
    ws.cell(row=r + 1, column=2, value="Filas respondidas")
    ws.cell(row=r + 1, column=3, value=0).fill = YELLOW
    ws.cell(row=r + 1, column=4, value=f"de {n_rows}")

    r += 3
    ws.cell(row=r, column=2, value="Comentarios sobre el instrumento").font = Font(size=12, bold=True)
    ws.cell(row=r + 1, column=2,
            value="¿Hubo algo confuso, mal redactado o difícil de decidir? Esto es lo más valioso.")
    ws.cell(row=r + 2, column=2).fill = YELLOW
    ws.merge_cells(start_row=r + 2, start_column=2, end_row=r + 6, end_column=5)

    r += 9
    ws.cell(row=r, column=2, value="Declaración").font = Font(size=12, bold=True)
    ws.cell(row=r + 1, column=2,
            value="Declaro que las respuestas de este instrumento reflejan mi criterio "
                  "profesional y que no consulté buscadores, herramientas de inteligencia "
                  "artificial ni a terceros para emitirlas.")
    ws.cell(row=r + 1, column=2).alignment = Alignment(wrap_text=True)

    r += 3
    ws.cell(row=r, column=2, value="Nombre").font = Font(bold=True)
    ws.cell(row=r, column=3, value="Colegiatura / credencial").font = Font(bold=True)
    ws.cell(row=r, column=4, value="Fecha").font = Font(bold=True)
    ws.cell(row=r, column=5, value="Conforme (Sí/No)").font = Font(bold=True)
    for col in (2, 3, 4, 5):
        ws.cell(row=r + 1, column=col).fill = YELLOW

    ws.cell(row=r + 3, column=2,
            value="Las celdas amarillas son las que se llenan. El resto es solo de lectura.")


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    parser.add_argument("--n-queries", type=int, default=MAX_N_QUERIES,
                         help=f"Número de consultas a usar ({MIN_N_QUERIES}-{MAX_N_QUERIES}). "
                              f"Decidir junto con la disponibilidad confirmada del adjudicador "
                              f"(MATCHING-SPEC.md §4.4.1), no antes.")
    parser.add_argument("--duplicate-rate", type=float, default=DEFAULT_DUPLICATE_RATE,
                         help=f"Proporción de pares duplicados, > 0.10 (default {DEFAULT_DUPLICATE_RATE}). "
                              f"El piloto validó que la marca de confianza baja predice "
                              f"inconsistencia intra-evaluador; subir la tasa de duplicados "
                              f"aumenta la potencia de esa medición.")
    parser.add_argument("--justification-rate", type=float, default=DEFAULT_JUSTIFICATION_RATE,
                         help=f"Proporción de filas con justificación escrita (default {DEFAULT_JUSTIFICATION_RATE}).")
    parser.add_argument("--seed", type=int, default=RANDOM_STATE)
    parser.add_argument("--out", type=Path,
                         default=Path(__file__).resolve().parents[1] / "data" / "reference" / "ds04_pool.xlsx")
    args = parser.parse_args()

    if args.duplicate_rate <= 0.10:
        raise SystemExit(
            f"--duplicate-rate debe ser > 0.10 (hallazgo del piloto: "
            f"10% dejaba la medición de consistencia intra-evaluador con muy poca potencia). "
            f"Recibido: {args.duplicate_rate}"
        )

    lawyers = json.loads(CORPUS_PATH.read_text(encoding="utf-8"))
    queries = select_queries(args.n_queries)
    rng = Random(args.seed)
    rows = build_rows(queries, lawyers, rng, args.duplicate_rate, args.justification_rate)

    n_duplicated_pairs = sum(1 for r in rows) - len({(r["lawyer_id"], r["case_description"]) for r in rows})
    print(f"Consultas: {len(queries)} de {MAX_N_QUERIES}")
    print(f"Filas totales: {len(rows)} (pares únicos: {len(rows) - n_duplicated_pairs}, "
          f"duplicados: {n_duplicated_pairs})")
    print(f"Filas con justificación: {sum(1 for r in rows if r['justify'])} "
          f"({sum(1 for r in rows if r['justify']) / len(rows) * 100:.1f}%)")

    wb = Workbook()
    ws_instr = wb.active
    ws_instr.title = "Instrucciones"
    write_instrucciones(ws_instr)

    ws_adj = wb.create_sheet("Adjudicación")
    write_adjudicacion(ws_adj, rows)

    ws_reg = wb.create_sheet("Registro")
    write_registro(ws_reg, len(rows))

    args.out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(args.out)
    print(f"Escrito: {args.out.resolve()}")


if __name__ == "__main__":
    main()
